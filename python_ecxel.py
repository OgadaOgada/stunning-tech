
from fileinput import filename
import re
from socket import herror
from traceback import print_tb
import openpyxl
import os
os.chdir(r"Q:\Learning\Python\stunning-tech") #change the working file path to this

# initial_excel = "openpyxl_file.xlsx"
# excelfile = openpyxl.load_workbook(initial_excel)
# print(type(ecxelfile))
# # sheet = ecxelfile.get_sheet_by_name("Sheet1") #deprecated
# sheet = excelfile["Sheet1"]
# # sheet = ecxelfile[ecxelfile.sheetnames[0]]
# # print(type(sheet))

# # READING FROM AN EXCEL WORKBOOK

# # sheet = ecxelfile.sheetnames
# print(sheet)
# # for i in sheet:
# #     print(i)
# # c2 = sheet["C2"] #passing the cell to a variable
# # print(c2.value) #getting the cell value
# cell_value = sheet["C2"].value
# print(cell_value)
# print(sheet.cell(row=2,column=18278)) #max colum according to this python

# # to print values in a single colum
# print("Name","Score")
# def cell_desc(cell_row,cell_col):
#     return sheet.cell(row=cell_row,column=cell_col).value #get the value of cell e.g cell A1 is cell(1,1)

# for i in range(2,9):
#     print(cell_desc(i,3),cell_desc(i,4))


# # CREATING AN EXEL WORKBOOK
# new_workbook = openpyxl.Workbook()
# score_sheet = new_workbook.create_sheet("Created")
# # new_sheet = new_workbook.sheetnames
# # del new_workbook["Sheet"]
# rename_it = new_workbook["Sheet"]
# rename_it.title = "Renamed"
# # print("sheets in new",new_sheet)
# print(score_sheet["A1"].value == None)
# score_sheet["A1"].value = "Marks".upper()
# score_sheet["A2"].value = 25
# score_sheet["A3"].value = 46
# score_sheet["B1"].value = "Name".upper()
# score_sheet["B2"].value = "Felicity Achieng"
# score_sheet["B3"].value = "Collince Ogada"
# for i in range(1,5):
#     if i == 1:
#         print(score_sheet.cell(row=i,column=1).value,score_sheet.cell(row=i,column=2).value)
#     else:
#         print(score_sheet.cell(row=i,column=1).value,score_sheet.cell(row=i,column=2).value)

new_excel_name = "new_excel.xlsx"
# new_sheet2 = new_workbook.create_sheet("Terms")
# new_sheet2 = new_workbook.create_sheet(index=0,title="Last Created Sheet")

# new_workbook.save(new_excel_name) #saved in the folder path specified up here by os.chdir()
# os.startfile(new_excel_name) #open the file after saving it




# COPY SHEET THREE[2] CONTENT TO SHEET ONE IN ANOTHER OPENPYXL.WORKBOOK

# # opening the source excel file sheet
# copied_excel = openpyxl.load_workbook(new_excel_name)
# sheet_copied = copied_excel.worksheets[2]
  
# # opening the destination excel file 
# last_excel_name = "last.xlsx"
# last_excel_wb = openpyxl.load_workbook(last_excel_name)
# sheet_pasted = last_excel_wb.active

  
# # calculate total number of rows and 
# # columns in source excel file
# myrows = sheet_copied.max_row
# mycols = sheet_copied.max_column
  
# # copying the cell values from source 
# # excel file to destination excel file
# for i in range (1, myrows + 1):
#     for j in range (1, mycols + 1):
#         # reading cell value from source excel file
#         c = sheet_copied.cell(row = i, column = j)
  
#         # writing the read value to destination excel file
#         sheet_pasted.cell(row = i, column = j).value = c.value

# last_excel_wb.save(last_excel_name) #saved in the folder path specified up here by os.chdir()
# os.startfile(last_excel_name) #open the file after saving it





# COPY SHEET ONE[0] CONTENT TO ALL SHEETS IN SAME OPENPYXL.WORKBOOK

# # opening the excel file 
# last_excel_name = "last.xlsx" #existing workbook file
# copy_to_all = openpyxl.load_workbook(last_excel_name)
# all_sheets = copy_to_all.sheetnames
# sheet_copied = copy_to_all.worksheets[0]
# # sheet_pasted = copy_to_all.worksheets[2]
  
# # calculate total number of rows and 
# # columns in source excel file
# myrows = sheet_copied.max_row
# mycols = sheet_copied.max_column
  
# # copying the cell values from source 
# # excel file to destination excel file
# for i in range (1, myrows + 1):
#     for j in range (1, mycols + 1):
#         # reading cell value from source excel file
#         source_cell_value = sheet_copied.cell(row = i, column = j).value
  
#         # writing the read value to destination excel file
#         for sheet in all_sheets:
#             if sheet != "Sheet1":
#                 # print(copy_to_all[sheet].cell(row = 1,column = 1).value)
#                 copy_to_all[sheet].cell(row = i, column = j).value = source_cell_value

# copy_to_all.save(last_excel_name) #saved in the folder path specified up here by os.chdir()
# os.startfile(last_excel_name) #open the file after saving it



more_sheets = "creating_several_worksheets.xlsx"
deleted_sheets = []
yes_options = ["Y","YES"]
no_options = ["N","NO"]
my_counter = 0
good_bye = "Good bye dude!"
many_attempts = "Too many attempts! Good bye."

def load_file(filename):
    return openpyxl.load_workbook(filename)

def save_file(filename,loaded_file):
    # pass
    return loaded_file.save(filename)

#delete everything except sheet at index 0
def Delete_all_sheets(filename):
    my_excel = load_file(filename)
    sheets = my_excel.sheetnames
    print(Show_sheets(filename))
    
    user_choice = input("Confirm to delete all sheets? (y/n)")
    if user_choice.upper == "Y":
        for sheet in sheets:
            if my_excel.worksheets.index(my_excel[sheet]) == 0:
                del my_excel[sheet]
                deleted_sheets.append(sheet)
        save_file(filename,my_excel)
        print("Sheets below were deleted successfully!")
        return "{}".format(deleted_sheets)
    else:
        return
    
#get all sheets from the file
def Show_sheets(filename):    
    my_excel = load_file(filename)
    return print(my_excel.sheetnames)

# user defined  choice to either delete by enetering sheet name or sheet index
def user_deletion_choice(filename):
    print("Available sheets: ",Show_sheets(filename))
    user_choice = input("Delete sheet by name or index (name/index)? ")
    if user_choice.lower() == "name":
        return False

    elif user_choice.lower() == "index":
        return True
    else:
        print("Invalid input")

#get user defined index to delete
def del_index(filename):
    my_excel = load_file(filename)  
    sheet_index = int(input("Enter the sheet index to delete: "))
    all_sheets_length = len(Show_sheets(filename))
    # if sheet_index > all_sheets_length or sheet_index  to check how to compare to negative index
    sheet_name = my_excel.sheetnames[sheet_index]
    confirm = input("Are you sure to delete {}? (y/n) ".format(sheet_name))
    if confirm.upper() == "Y":
        del my_excel[sheet_name]
        save_file(filename,my_excel)
    else:
        return

#get user defined sheet name to delete
def del_name(filename):
    my_excel = load_file(filename)  
     # to check if sheet name exist first
    sheet_name = input("Enter the sheet name to delete: ")
    if sheet_name in Show_sheets(filename):
    # sheet_name = my_excel.sheetnames[sheet_index]
        confirm = input("Are you sure to delete {}? (y/n) ".format(sheet_name))
        if confirm.upper() == "Y":
            del my_excel[sheet_name]
            save_file(filename,my_excel)
        else:
            return
    else:
        print("sheet name does not exist!")

#perfom the specific delete operation
def Delete_specific_sheet(filename):
    user_input = user_deletion_choice(filename)
    if user_input: #returned true, delete by index
        del_index(filename)
    elif not user_input: #returned false, delete by name
        del_name(filename)
    return

# create user defined number of sheets
def Create_sheets(filename):
    print(Show_sheets(filename))
    num_sheets = int(input("How many sheets to create: "))

    #create sheets on an existing workbook
    my_excel = load_file(filename)
    if num_sheets == 1:        
        my_excel.create_sheet("Sheet"+str(num_sheets))
        save_file(filename,my_excel)
        print(Show_sheets(filename))

    elif num_sheets>0:
        for i in range(1,num_sheets+1):
            my_excel.create_sheet("Sheet"+str(i))
        save_file(filename,my_excel)
        print(Show_sheets(filename))
    else:
        print("Invalid number!")
    return

def Copy_to_sheet(filename):
    # how to copy to sheet name, get index from sheet name
    # have two options, copy by index or sheet name
    my_excel = load_file(filename)
    print(Show_sheets(filename))
    copy_from = int(input("Index of sheet to copy: "))
    copy_to = int(input("Index of sheet to paste to: "))
    sheet_copied = my_excel.worksheets[copy_from]
    sheet_pasted = my_excel.worksheets[copy_to]
    from_name = my_excel.sheetnames[copy_from]
    to_name = my_excel.sheetnames[copy_to]
    
    # calculate total number of rows and columns in source excel file
    myrows = sheet_copied.max_row
    mycols = sheet_copied.max_column
    
    # copying the cell values from source excel file to destination excel file
    for i in range (1, myrows + 1):
        for j in range (1, mycols + 1):
            # reading cell value from source excel file
            source_cell_value = sheet_copied.cell(row = i, column = j).value
    
            # writing the read value to destination excel file
            sheet_pasted.cell(row = i, column = j).value = source_cell_value
    save_file(filename,my_excel)
    print("Copied succefully from {} to {}.".format(from_name,to_name))


def Open_the_file(filename):
    # filename.WindowState = win32.constants.xlMaximized
    # to research more on maximizing the windows after start
    return os.startfile(filename)


def another_operation(filename):
    another_action = input("\nPerform another action? (y/n): ").upper()
    global my_counter #to prevent overwriting the value
    
    if another_action in yes_options:
        my_counter = 0
        return my_menu(filename)

    elif another_action in no_options:
        my_counter = 0
        print(good_bye)
        return quit()
    else:
        my_counter +=1
        if my_counter == 3:
            print(many_attempts)
            quit()
        else:
            print("\nInvalid choice! {} attempt(s) remaining.".format(3 - my_counter))
            return another_operation(filename)
    
def Rename_sheet():
    pass
def Enter_text():
    pass
def Search_pattern():
    pass

ops_dictionary = {"Show_sheets":Show_sheets,"Create_sheet":Create_sheets, "Delete_all_sheets":Delete_all_sheets, "Delete_specific_sheet":Delete_specific_sheet,
                     "Copy_to_sheet":Copy_to_sheet, "Rename_sheet":Rename_sheet,"Enter_text":Enter_text, "Open_the_file":Open_the_file,
                     "Search_pattern":Search_pattern}
ops_list = list(ops_dictionary)

#display the actions menu for the user
def my_menu(filename):
    print()
    print("Operations Menu".upper())
    for i in ops_list:
        print(ops_list.index(i)+1,i)
    print("_________________________________")
    print("The excel file: '{}'\n".format(filename))
    user_selection(filename)
    return 

def func_call(func,filename):
    try:
        return ops_dictionary[func](filename)

    except Exception as err:
        print("Invalid selection", err)

def user_selection(filename):
    selection = int(input("Selection: "))
    my_function = ops_list[selection - 1] #getting the function name from a list using the entered value index
    func_call(my_function,filename)
    # print(my_function)
    another_operation(filename)


print("TO WORK ON HOW TO CLOSE OPENED EXCEL FILES")
# print("CHANGE THE FUNCTION NAME TO DICTIONARY FOR EASY CREATION") done
# another_operation(more_sheets)
print()
my_menu(more_sheets)
# user_selection(filename)
